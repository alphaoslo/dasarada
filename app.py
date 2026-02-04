from flask import Flask, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "ticket_data.xlsx"

# Create Excel file with headers if it does not exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Name", "Email", "Phone Number", "Age", "Gender",
        "From City", "To City", "Journey Date",
        "Tickets", "Additional Details"
    ])
    wb.save(EXCEL_FILE)

@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name")
    email = request.form.get("email")
    phone = request.form.get("phonenumber")
    age = request.form.get("age")
    gender = request.form.get("gender")
    from_city = request.form.get("fromcityname")
    to_city = request.form.get("tocityname")
    journey_date = request.form.get("journey_date")
    tickets = request.form.get("tickets")
    details = request.form.get("details")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([
        name, email, phone, age, gender,
        from_city, to_city, journey_date,
        tickets, details
    ])

    wb.save(EXCEL_FILE)

    return "<h2>Ticket booked successfully!</h2>"

if __name__ == "__main__":
    app.run(debug=True)
